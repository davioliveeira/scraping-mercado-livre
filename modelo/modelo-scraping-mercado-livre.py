import datetime
import openpyxl
from openpyxl.styles.numbers import NumberFormatDescriptor # IMPORTA A FORMATAÇÃO DE DATA E HORARIO  #
from time import sleep
from selenium import webdriver # IMPORTA O NAVEGADOR # 
from selenium.webdriver.chrome.options import Options #  POSSIBILITA PASSAR CONFIGURAÇOS DO BROWSER  #
from selenium.webdriver.common.keys import Keys # POSSIBILITA A UTILIZAÇÃO DE KEYS #
from selenium.webdriver.common.by import By # POSSIBILITA A BUSCA DE ELEMENTOS HTML PELO NAVEGADOR #

#  DEFINIÇÃO DE ESCOPO  #
busca_produto = 'Macbook Pro'
url = 'https://www.mercadolivre.com.br/'

#  MANIPULA A DATA E FORMATA ELA PARA O PADRAO dd/mm/aaaa  #
data_atual = datetime.datetime.now().strftime("%d/%m/%Y")

#  DEFINE ALGUMAS CONFIGURAÇÕES PARA O BROWSER  #
options = Options()
options.add_argument("--start-maximized")

#  INICIA O BROWSER ESCOLHIDO PASSANDO AS CONFIGURAÇÕES SETADAS  #
driver = webdriver.Chrome(options=options)

#  ABRE O BROWSER E ACESSA A URL DO MERCADO LIVRE  # 
driver.get(url)

# FAZ UMA ASSERTIVA VERIFICANDO O TITULO DO SITE É "Mercado Livre"  #
assert "Mercado Livre" in driver.title

#  PROCURA O ELEMENTO DE BUSCA DO SITE E REALIZA A PESQUISA  #
campo_busca = driver.find_element(By.NAME, 'as_word')
campo_busca.send_keys(busca_produto)
campo_busca.send_keys(Keys.RETURN)

#  AGUARDA O RESULTADO DA PESQUISA CARREGAR NO BROWSER  #
driver.implicitly_wait(10)

#  PROCURA PELOS PRODUTOS DA LISTADOS NO BROWSER  # 
produtos = driver.find_elements(By.CSS_SELECTOR, '.ui-search-layout__item')

#  CRIA UM ARQUIVO EXCEL  #
planilha = openpyxl.Workbook()
aba = planilha.active

# NOMEIA AS COLUNAS DO EXCEL  # 
aba['A1'] = 'Nome do Produto'
aba['B1'] = 'Preço'
aba['C1'] = 'Data de Extração'

#  FAZ UMA MANIPULAÇAO DA VARIAVEL BUSCA PRODUTO PARA UMA MELHOR NOMEAÇAO DO ARQUIVO EXCEL POSTERIORMENTE  # 
produto = busca_produto

#  FAZ UM LOOP ITERANDO OS PRODUTOS E INSERE NO ARQUIVO EXCEL  #
for linha, produto in enumerate(produtos, start=2):
    #  ABRE UM BLOCO DE TRY CAT (BASICAMENTE TRATA EXCEÇÕES EXISTENTES)  #
    try:
        #  EXTRAI O NOME E O PREÇO DO PRODUTO  # 
        nome_produto = produto.find_element(By.CSS_SELECTOR, '.ui-search-item__title').text
        preco = produto.find_element(By.CSS_SELECTOR, '.price-tag-fraction').text
        
        #  INSERE OS DADOS EXTRAIDOS NAS RESPECTIVAS COLUNAS DO EXCEL  #
        aba.cell(row=linha, column=1, value=nome_produto)
        aba.cell(row=linha, column=2, value=preco)
        aba.cell(row=linha, column=3, value=data_atual)

        #  IMPRIME O RESULTADO PROCESSADO PARA CADA PRODUTO  #
        print(f"Processando produto: {nome_produto}, preço: {preco}")

    #  TRATA E APRESENTA ALGUMA EXCEÇÃO/ERROR QUE VENHA A APARECER  #    
    except Exception as e:
        print(f"Erro ao processar produto: {e}")


#  SALVA A PLANILHA CRIADA  # 
nome_arquivo = f'{busca_produto}_{datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
planilha.save(nome_arquivo)

#  FECHA O NAVEGADOR  #
driver.quit()
